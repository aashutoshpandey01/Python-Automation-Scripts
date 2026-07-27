import csv
import json
import platform
import socket
import sqlite3
import subprocess
import sys

from datetime import datetime, timezone
from pathlib import Path


# ============================================================
# BASIC INFORMATION
# ============================================================

HOSTNAME = platform.node()

OPERATING_SYSTEM = platform.system()

TIMESTAMP = datetime.now(
    timezone.utc
).isoformat()


# ============================================================
# AUTOMATIC DIRECTORIES
# ============================================================

BASE_DIRECTORY = Path(__file__).parent

DATA_DIRECTORY = (
    BASE_DIRECTORY
    / "inventory_asset_data"
    / HOSTNAME
)

REPORT_DIRECTORY = DATA_DIRECTORY / "reports"

REPORT_DIRECTORY.mkdir(
    parents=True,
    exist_ok=True
)


# ============================================================
# CPU INFORMATION
# ============================================================

def collect_cpu_information():

    return {

        "processor": platform.processor(),

        "cpu_architecture": platform.machine(),

        "logical_cpu_count": (
            __import__("os").cpu_count()
        )

    }


# ============================================================
# MEMORY INFORMATION
# ============================================================

def collect_memory_information():

    try:

        import psutil

        memory = psutil.virtual_memory()

        return {

            "total_gb": round(
                memory.total / (1024 ** 3),
                2
            ),

            "available_gb": round(
                memory.available / (1024 ** 3),
                2
            ),

            "used_gb": round(
                memory.used / (1024 ** 3),
                2
            ),

            "usage_percent": memory.percent

        }

    except ImportError:

        return {

            "status": (
                "psutil is not installed"
            ),

            "message": (
                "Install psutil for memory "
                "information"
            )

        }


# ============================================================
# DISK INFORMATION
# ============================================================

def collect_disk_information():

    disks = []

    try:

        import psutil

        partitions = psutil.disk_partitions(
            all=False
        )


        for partition in partitions:

            try:

                usage = psutil.disk_usage(
                    partition.mountpoint
                )


                disks.append({

                    "device": partition.device,

                    "mountpoint": (
                        partition.mountpoint
                    ),

                    "filesystem": (
                        partition.fstype
                    ),

                    "total_gb": round(
                        usage.total
                        / (1024 ** 3),
                        2
                    ),

                    "used_gb": round(
                        usage.used
                        / (1024 ** 3),
                        2
                    ),

                    "free_gb": round(
                        usage.free
                        / (1024 ** 3),
                        2
                    ),

                    "usage_percent": (
                        usage.percent
                    )

                })


            except Exception:

                continue


    except ImportError:

        return {

            "status": (
                "psutil is not installed"
            )

        }


    return disks


# ============================================================
# IP ADDRESS INFORMATION
# ============================================================

def collect_ip_addresses():

    ip_addresses = []

    try:

        hostname = socket.gethostname()


        addresses = socket.getaddrinfo(

            hostname,

            None

        )


        for address in addresses:

            ip_address = address[4][0]


            if ip_address not in ip_addresses:

                ip_addresses.append(
                    ip_address
                )


    except Exception as error:

        return {

            "error": str(error)

        }


    return ip_addresses


# ============================================================
# INSTALLED SOFTWARE
# ============================================================

def collect_installed_software():

    operating_system = platform.system()


    software = []


    # --------------------------------------------------------
    # WINDOWS
    # --------------------------------------------------------

    if operating_system == "Windows":

        try:

            command = [

                "powershell",

                "-NoProfile",

                "-Command",

                (

                    "Get-ItemProperty "

                    "HKLM:\\Software\\Microsoft\\"

                    "Windows\\CurrentVersion\\"

                    "Uninstall\\* "

                    "| Where-Object "

                    "{$_.DisplayName} "

                    "| Select-Object "

                    "DisplayName, DisplayVersion "

                    "| ConvertTo-Json"

                )

            ]


            result = subprocess.run(

                command,

                capture_output=True,

                text=True,

                timeout=60

            )


            if result.stdout.strip():

                software_data = json.loads(

                    result.stdout

                )


                if isinstance(

                    software_data,

                    dict

                ):

                    software_data = [

                        software_data

                    ]


                for item in software_data:

                    software.append({

                        "name": item.get(

                            "DisplayName"

                        ),

                        "version": item.get(

                            "DisplayVersion"

                        )

                    })


        except Exception as error:

            software.append({

                "error": str(error)

            })


    # --------------------------------------------------------
    # LINUX
    # --------------------------------------------------------

    elif operating_system == "Linux":

        try:

            result = subprocess.run(

                [

                    "dpkg",

                    "-l"

                ],

                capture_output=True,

                text=True,

                timeout=60

            )


            for line in result.stdout.splitlines():

                if line.startswith("ii"):

                    parts = line.split()


                    if len(parts) >= 3:

                        software.append({

                            "name": parts[1],

                            "version": parts[2]

                        })


        except FileNotFoundError:

            try:

                result = subprocess.run(

                    [

                        "rpm",

                        "-qa"

                    ],

                    capture_output=True,

                    text=True,

                    timeout=60

                )


                for line in (

                    result.stdout.splitlines()

                ):

                    software.append({

                        "name": line

                    })


            except Exception as error:

                software.append({

                    "error": str(error)

                })


        except Exception as error:

            software.append({

                "error": str(error)

            })


    else:

        software.append({

            "status": (

                "Operating system not supported"

            )

        })


    return software


# ============================================================
# RUNNING SERVICES
# ============================================================

def collect_running_services():

    operating_system = platform.system()

    services = []


    # --------------------------------------------------------
    # WINDOWS
    # --------------------------------------------------------

    if operating_system == "Windows":

        try:

            result = subprocess.run(

                [

                    "powershell",

                    "-NoProfile",

                    "-Command",

                    (

                        "Get-Service "

                        "| Select-Object "

                        "Name,DisplayName,Status "

                        "| ConvertTo-Json"

                    )

                ],

                capture_output=True,

                text=True,

                timeout=60

            )


            if result.stdout.strip():

                service_data = json.loads(

                    result.stdout

                )


                if isinstance(

                    service_data,

                    dict

                ):

                    service_data = [

                        service_data

                    ]


                for service in service_data:

                    services.append({

                        "name": service.get(

                            "Name"

                        ),

                        "display_name": service.get(

                            "DisplayName"

                        ),

                        "status": service.get(

                            "Status"

                        )

                    })


        except Exception as error:

            services.append({

                "error": str(error)

            })


    # --------------------------------------------------------
    # LINUX
    # --------------------------------------------------------

    elif operating_system == "Linux":

        try:

            result = subprocess.run(

                [

                    "systemctl",

                    "list-units",

                    "--type=service",

                    "--all",

                    "--no-pager",

                    "--no-legend"

                ],

                capture_output=True,

                text=True,

                timeout=60

            )


            for line in result.stdout.splitlines():

                parts = line.split()

                if len(parts) >= 4:

                    services.append({

                        "name": parts[0],

                        "load": parts[1],

                        "active": parts[2],

                        "sub": parts[3]

                    })


        except Exception as error:

            services.append({

                "error": str(error)

            })


    else:

        services.append({

            "status": (

                "Operating system not supported"

            )

        })


    return services


# ============================================================
# COLLECT ALL INVENTORY
# ============================================================

def collect_inventory():

    inventory = {

        "collection_time": TIMESTAMP,

        "hostname": HOSTNAME,

        "operating_system": {

            "name": OPERATING_SYSTEM,

            "release": platform.release(),

            "version": platform.version(),

            "architecture": (
                platform.architecture()[0]
            )

        },

        "cpu": (

            collect_cpu_information()

        ),

        "memory": (

            collect_memory_information()

        ),

        "disks": (

            collect_disk_information()

        ),

        "ip_addresses": (

            collect_ip_addresses()

        ),

        "installed_software": (

            collect_installed_software()

        ),

        "running_services": (

            collect_running_services()

        )

    }


    return inventory


# ============================================================
# SAVE JSON REPORT
# ============================================================

def save_json_report(inventory):

    timestamp = datetime.now().strftime(

        "%Y%m%d_%H%M%S"

    )


    json_file = (

        REPORT_DIRECTORY

        / f"inventory_{timestamp}.json"

    )


    with open(

        json_file,

        "w",

        encoding="utf-8"

    ) as file:

        json.dump(

            inventory,

            file,

            indent=4

        )


    return json_file


# ============================================================
# SAVE CSV REPORT
# ============================================================

def save_csv_report(inventory):

    timestamp = datetime.now().strftime(

        "%Y%m%d_%H%M%S"

    )


    csv_file = (

        REPORT_DIRECTORY

        / f"inventory_{timestamp}.csv"

    )


    rows = []


    # --------------------------------------------------------
    # SYSTEM
    # --------------------------------------------------------

    rows.append({

        "category": "System",

        "name": "Hostname",

        "value": inventory["hostname"]

    })


    rows.append({

        "category": "System",

        "name": "Operating System",

        "value": inventory[
            "operating_system"
        ]["name"]

    })


    rows.append({

        "category": "System",

        "name": "OS Version",

        "value": inventory[
            "operating_system"
        ]["version"]

    })


    # --------------------------------------------------------
    # CPU
    # --------------------------------------------------------

    for key, value in (

        inventory["cpu"].items()

    ):

        rows.append({

            "category": "CPU",

            "name": key,

            "value": value

        })


    # --------------------------------------------------------
    # MEMORY
    # --------------------------------------------------------

    for key, value in (

        inventory["memory"].items()

    ):

        rows.append({

            "category": "Memory",

            "name": key,

            "value": value

        })


    # --------------------------------------------------------
    # IP ADDRESSES
    # --------------------------------------------------------

    for ip_address in (

        inventory["ip_addresses"]

    ):

        rows.append({

            "category": "Network",

            "name": "IP Address",

            "value": ip_address

        })


    # --------------------------------------------------------
    # DISKS
    # --------------------------------------------------------

    for disk in inventory["disks"]:

        rows.append({

            "category": "Disk",

            "name": disk.get(

                "mountpoint"

            ),

            "value": json.dumps(

                disk

            )

        })


    # --------------------------------------------------------
    # SOFTWARE
    # --------------------------------------------------------

    for software in (

        inventory["installed_software"]

    ):

        rows.append({

            "category": "Software",

            "name": software.get(

                "name"

            ),

            "value": software.get(

                "version"

            )

        })


    # --------------------------------------------------------
    # SERVICES
    # --------------------------------------------------------

    for service in (

        inventory["running_services"]

    ):

        rows.append({

            "category": "Service",

            "name": service.get(

                "name"

            ),

            "value": json.dumps(

                service

            )

        })


    with open(

        csv_file,

        "w",

        newline="",

        encoding="utf-8"

    ) as file:

        writer = csv.DictWriter(

            file,

            fieldnames=[

                "category",

                "name",

                "value"

            ]

        )


        writer.writeheader()


        writer.writerows(rows)


    return csv_file


# ============================================================
# SAVE EXCEL REPORT
# ============================================================

def save_excel_report(inventory):

    try:

        from openpyxl import Workbook


    except ImportError:

        return (

            "Excel skipped: "

            "openpyxl is not installed"

        )


    timestamp = datetime.now().strftime(

        "%Y%m%d_%H%M%S"

    )


    excel_file = (

        REPORT_DIRECTORY

        / f"inventory_{timestamp}.xlsx"

    )


    workbook = Workbook()


    # --------------------------------------------------------
    # SYSTEM SHEET
    # --------------------------------------------------------

    system_sheet = (

        workbook.active

    )

    system_sheet.title = "System"


    system_sheet.append([

        "Property",

        "Value"

    ])


    system_sheet.append([

        "Hostname",

        inventory["hostname"]

    ])


    system_sheet.append([

        "Operating System",

        inventory[

            "operating_system"

        ]["name"]

    ])


    system_sheet.append([

        "OS Version",

        inventory[

            "operating_system"

        ]["version"]

    ])


    # --------------------------------------------------------
    # CPU SHEET
    # --------------------------------------------------------

    cpu_sheet = workbook.create_sheet(

        "CPU"

    )


    cpu_sheet.append([

        "Property",

        "Value"

    ])


    for key, value in (

        inventory["cpu"].items()

    ):

        cpu_sheet.append([

            key,

            value

        ])


    # --------------------------------------------------------
    # MEMORY SHEET
    # --------------------------------------------------------

    memory_sheet = workbook.create_sheet(

        "Memory"

    )


    memory_sheet.append([

        "Property",

        "Value"

    ])


    for key, value in (

        inventory["memory"].items()

    ):

        memory_sheet.append([

            key,

            value

        ])


    # --------------------------------------------------------
    # DISKS SHEET
    # --------------------------------------------------------

    disk_sheet = workbook.create_sheet(

        "Disks"

    )


    disk_sheet.append([

        "Device",

        "Mountpoint",

        "Filesystem",

        "Total GB",

        "Used GB",

        "Free GB",

        "Usage %"

    ])


    for disk in inventory["disks"]:

        disk_sheet.append([

            disk.get("device"),

            disk.get("mountpoint"),

            disk.get("filesystem"),

            disk.get("total_gb"),

            disk.get("used_gb"),

            disk.get("free_gb"),

            disk.get("usage_percent")

        ])


    # --------------------------------------------------------
    # IP ADDRESSES SHEET
    # --------------------------------------------------------

    network_sheet = workbook.create_sheet(

        "Network"

    )


    network_sheet.append([

        "IP Address"

    ])


    for ip_address in (

        inventory["ip_addresses"]

    ):

        network_sheet.append([

            ip_address

        ])


    # --------------------------------------------------------
    # SOFTWARE SHEET
    # --------------------------------------------------------

    software_sheet = workbook.create_sheet(

        "Software"

    )


    software_sheet.append([

        "Software",

        "Version"

    ])


    for software in (

        inventory["installed_software"]

    ):

        software_sheet.append([

            software.get("name"),

            software.get("version")

        ])


    # --------------------------------------------------------
    # SERVICES SHEET
    # --------------------------------------------------------

    services_sheet = workbook.create_sheet(

        "Services"

    )


    services_sheet.append([

        "Service",

        "Status"

    ])


    for service in (

        inventory["running_services"]

    ):

        services_sheet.append([

            service.get("name"),

            service.get(

                "status",

                service.get(

                    "active"

                )

            )

        ])


    workbook.save(

        excel_file

    )


    return excel_file


# ============================================================
# SAVE SQLITE DATABASE
# ============================================================

def save_database(inventory):

    timestamp = datetime.now().strftime(

        "%Y%m%d_%H%M%S"

    )


    database_file = (

        REPORT_DIRECTORY

        / "inventory.db"

    )


    connection = sqlite3.connect(

        database_file

    )


    cursor = connection.cursor()


    # --------------------------------------------------------
    # SYSTEM TABLE
    # --------------------------------------------------------

    cursor.execute(

        """

        CREATE TABLE IF NOT EXISTS system_inventory (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            collection_time TEXT,

            hostname TEXT,

            operating_system TEXT,

            os_version TEXT

        )

        """

    )


    cursor.execute(

        """

        INSERT INTO system_inventory (

            collection_time,

            hostname,

            operating_system,

            os_version

        )

        VALUES (?, ?, ?, ?)

        """,

        (

            inventory["collection_time"],

            inventory["hostname"],

            inventory[

                "operating_system"

            ]["name"],

            inventory[

                "operating_system"

            ]["version"]

        )

    )


    # --------------------------------------------------------
    # SOFTWARE TABLE
    # --------------------------------------------------------

    cursor.execute(

        """

        CREATE TABLE IF NOT EXISTS installed_software (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            collection_time TEXT,

            hostname TEXT,

            software_name TEXT,

            version TEXT

        )

        """

    )


    for software in (

        inventory["installed_software"]

    ):

        cursor.execute(

            """

            INSERT INTO installed_software (

                collection_time,

                hostname,

                software_name,

                version

            )

            VALUES (?, ?, ?, ?)

            """,

            (

                inventory["collection_time"],

                inventory["hostname"],

                software.get("name"),

                software.get("version")

            )

        )


    # --------------------------------------------------------
    # SERVICES TABLE
    # --------------------------------------------------------

    cursor.execute(

        """

        CREATE TABLE IF NOT EXISTS services (

            id INTEGER PRIMARY KEY AUTOINCREMENT,

            collection_time TEXT,

            hostname TEXT,

            service_name TEXT,

            status TEXT

        )

        """

    )


    for service in (

        inventory["running_services"]

    ):

        cursor.execute(

            """

            INSERT INTO services (

                collection_time,

                hostname,

                service_name,

                status

            )

            VALUES (?, ?, ?, ?)

            """,

            (

                inventory["collection_time"],

                inventory["hostname"],

                service.get("name"),

                service.get(

                    "status",

                    service.get(

                        "active"

                    )

                )

            )

        )


    connection.commit()

    connection.close()


    return database_file


# ============================================================
# MAIN
# ============================================================

def main():

    print(

        "Inventory and Asset Collection"

    )


    print(

        "Operating System:",

        OPERATING_SYSTEM

    )


    print(

        "Hostname:",

        HOSTNAME

    )


    print(

        "\nCollecting system information..."

    )


    inventory = collect_inventory()


    print(

        "Collecting CPU information..."

    )


    print(

        "Collecting memory information..."

    )


    print(

        "Collecting disk information..."

    )


    print(

        "Collecting IP addresses..."

    )


    print(

        "Collecting installed software..."

    )


    print(

        "Collecting running services..."

    )


    print(

        "\nSaving reports..."

    )


    json_file = save_json_report(

        inventory

    )


    csv_file = save_csv_report(

        inventory

    )


    excel_file = save_excel_report(

        inventory

    )


    database_file = save_database(

        inventory

    )


    print(

        "\nInventory collection completed."

    )


    print(

        "\nJSON report:"

    )

    print(

        json_file

    )


    print(

        "\nCSV report:"

    )

    print(

        csv_file

    )


    print(

        "\nExcel report:"

    )

    print(

        excel_file

    )


    print(

        "\nDatabase:"

    )

    print(

        database_file

    )


# ============================================================
# START
# ============================================================

if __name__ == "__main__":

    try:

        main()


    except Exception as error:

        print(

            "\nERROR:",

            error

        )